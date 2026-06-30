#!/usr/bin/env python3
"""Cross-zonal causation-vs-burden analysis from official PJM data.
Causation = zone large-load (data-centre) adjustment / RTO total (Table B-9b).
Burden    = zone annual peak / RTO annual peak (2026 Load Report data).
Both are official PJM figures; the divergence is the novel cross-zonal result.
"""
import openpyxl, csv, os
HERE = os.path.dirname(os.path.abspath(__file__))  # generated CSV written here

def _find_file(filename, candidates):
    """Return the first existing path from a list of candidates."""
    for p in candidates:
        if os.path.exists(p):
            return p
    raise FileNotFoundError(f"{filename} not found; tried: {candidates}")

# Table B-9b Excel — bundled in ../data/ in the released package
B9B_EXCEL = _find_file("total_load_adjustments_breakdown.xlsx", [
    os.path.join(HERE, "..", "data", "total_load_adjustments_breakdown.xlsx"),
    os.path.join(HERE, "pjm_extra", "total_load_adjustments_breakdown.xlsx"),
    os.path.join(HERE, "total_load_adjustments_breakdown.xlsx"),
])
LOAD_EXCEL = os.path.join(HERE, "pjm_2026_load_report_data.xlsx")
# NOTE: pjm_2026_load_report_data.xlsx is the per-zone Load Report data workbook
# from PJM 2026 Load Report (ref [16]; downloadable from pjm.com). Not bundled here
# because it is large; the pre-computed output crosszonal_caus_burden.csv IS included.

# --- Table B-9b: per-zone large-load adjustment (MW) by year ---
wb=openpyxl.load_workbook(os.path.join(HERE, B9B_EXCEL), data_only=True)
ws=wb["Total LargeLoad Breakdown"]; rows=list(ws.iter_rows(values_only=True)); hdr=rows[3]
yr_col={hdr[i]:i for i in range(len(hdr)) if isinstance(hdr[i],int)}
adj={}
for r in rows[4:]:
    z=r[0]
    if z and str(z).strip() and str(z).strip()!='\xa0' and r[3] not in (None,''):
        zn=str(z).strip()
        adj[zn]={y:(r[c] if r[c] is not None else 0) for y,c in yr_col.items()}
RTO_adj=adj.pop("PJM RTO")

# --- Load report: per-zone annual peak (MW) ---
wb2=openpyxl.load_workbook(os.path.join(HERE, LOAD_EXCEL), data_only=True)
peak={}
for sn in wb2.sheetnames:
    wsp=wb2[sn]; rr=list(wsp.iter_rows(values_only=True))
    if not rr or rr[0][0]!='ZONE_NAME': continue
    ann={}
    for r in rr[1:]:
        z,y,mo,pk = r[0],r[1],r[2],r[3]
        if y is None or pk is None: continue
        y=int(y); ann[y]=max(ann.get(y,0),float(pk))
    peak[sn]=ann
RTO_peak=peak["PJM_RTO"]

# zone-name map (B-9b -> peak sheet)
NAME={"DAYTON":"DAY","JCPL":"JCPL_FE_EAST","METED":"METED_FE_EAST","PENLC":"PN_FE_EAST"}
def peak_of(zn,y):
    k=NAME.get(zn,zn)
    return peak.get(k,{}).get(y)

YEARS=[2026,2030,2046]
out=[["zone"]+sum([[f"adj{y}",f"peak{y}",f"caus_share{y}_pct",f"burden_share{y}_pct",f"ratio{y}"] for y in YEARS],[])]
zones=[z for z in adj if peak_of(z,2026)]
for z in sorted(zones, key=lambda z:-adj[z][2026]):
    row=[z]
    for y in YEARS:
        a=adj[z].get(y,0); pk=peak_of(z,y)
        cs=100*a/RTO_adj[y] if RTO_adj[y] else 0
        bs=100*pk/RTO_peak[y] if pk else 0
        ratio=(cs/bs) if bs else 0
        row+=[a, pk, round(cs,2), round(bs,2), round(ratio,2)]
    out.append(row)
with open(os.path.join(HERE,"crosszonal_caus_burden.csv"),"w",newline="",encoding="utf-8") as f:
    csv.writer(f).writerows(out)

# summary stats for 2026
print(f"RTO 2026: total large-load adj = {RTO_adj[2026]:,} MW ; RTO peak = {RTO_peak[2026]:,.0f} MW")
print(f"{'zone':8s} {'adj26':>7s} {'caus%':>6s} {'burden%':>7s} {'ratio':>6s}")
import statistics as st
data=[]
for z in sorted(zones, key=lambda z:-adj[z][2026]):
    a=adj[z][2026]; pk=peak_of(z,2026)
    cs=100*a/RTO_adj[2026]; bs=100*pk/RTO_peak[2026]; ratio=cs/bs if bs else 0
    data.append((z,a,cs,bs,ratio))
    if a>0: print(f"{z:8s} {a:7.0f} {cs:6.1f} {bs:7.1f} {ratio:6.2f}")
# concentration: how many zones make up 80% of adjustment?
tot=RTO_adj[2026]; cum=0; n=0
for z,a,cs,bs,r in data:
    cum+=a; n+=1
    if cum>=0.8*tot: break
print(f"\nTop {n} zones (of {len(data)}) hold 80% of the 2026 large-load adjustment.")
top3=sorted(data,key=lambda x:-x[1])[:3]
print(f"Top 3 (DOM,AEP,COMED) causation share = {sum(x[2] for x in top3):.1f}% but burden share = {sum(x[3] for x in top3):.1f}%")
# Gini-like concentration of adjustment vs peak
print(f"DOM: causation {data[0][2]:.1f}% vs burden {data[0][3]:.1f}% -> ratio {data[0][4]:.2f}x")
# count zones that net-subsidize (burden>causation, i.e. ratio<1) among those with load
sub=[z for z,a,cs,bs,r in data if r<1]
drv=[z for z,a,cs,bs,r in data if r>1.2]
print(f"Zones with causation>burden (drivers, ratio>1.2): {drv}")
print(f"# zones causation<burden (net payers, ratio<1): {len(sub)} of {len(data)}")
